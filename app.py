from flask import Flask, render_template, request, jsonify, send_file, redirect
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pipeline import *
from pipeline import (
    DATA_CAGR_SVP, DATA_MARKET_SIZE, DATA_ALL_PRODUCTS, DATA_MARKET_HIERARCHIES,
    DATA_REVENUE_FR, DATA_REVENUE_INTL,
    build_category_mapper, combine_revenue_files,
)
import json
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
import pandas as pd

app = Flask(__name__)

revenue_years     = []
market_size_years = []
df_all_data       = None

TRANSLATIONS_DIR = os.path.join(BASE_DIR, 'translations')

def load_translations(language='en'):
    """Load translation file for given language"""
    try:
        filepath = os.path.join(TRANSLATIONS_DIR, f'{language}.json')
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Error loading translations: {e}")
    return {}

@app.route('/api/translations/<language>')
def get_translations(language):
    """API endpoint to fetch translations"""
    if language not in ['en', 'fr']:
        language = 'en'
    
    translations = load_translations(language)
    return jsonify(translations)

@app.route('/api/settings', methods=['GET', 'POST'])
def settings_api():
    """Get or update user settings"""
    if request.method == 'GET':
        # Return current settings from session/localStorage
        return jsonify({
            'language': request.args.get('language', 'en'),
            'theme': request.args.get('theme', 'dark-mode')
        })
    
    elif request.method == 'POST':
        data = request.json
        # Settings are stored in localStorage on client side
        return jsonify({'success': True})

@app.route('/')
def index():
    global revenue_years, market_size_years, df_all_data
    if not revenue_years:     revenue_years     = load_revenue_data()
    if not market_size_years: market_size_years = load_market_size_years()
    if df_all_data is None:   df_all_data       = load_all_data()

    svp_options = [
        '1.SVP - Employee Experience',
        '2.SVP - Digital Infrastructure',
        '3.SVP - Operational Experience',
        '4.SVP - Customer Experience',
    ]
    regions = ['France', 'International']

    default_display_start = revenue_years[0]      if revenue_years      else (market_size_years[0]  if market_size_years else 2020)
    default_display_end   = market_size_years[-1]  if market_size_years  else (revenue_years[-1]     if revenue_years     else 2030)
    default_cagr_start    = revenue_years[-1]      if revenue_years      else (market_size_years[0]  if market_size_years else 2023)
    default_cagr_end      = market_size_years[-1]  if market_size_years  else (revenue_years[-1]     if revenue_years     else 2030)
    default_pred_start    = revenue_years[-1]      if revenue_years      else (market_size_years[0]  if market_size_years else 2023)

    warnings = []
    if df_all_data is not None and 'missing_offers' in df_all_data.attrs:
        missing_offers = df_all_data.attrs['missing_offers']
        if missing_offers:
            warnings.append({
                'type': 'warning',
                'message': f'{len(missing_offers)} offers have incomplete data in CAGR Mapper',
                'details': missing_offers[:10]
            })

    # CVP options for the sidebar
    cvp_options = []
    if df_all_data is not None and 'cvp_options' in df_all_data.attrs:
        cvp_options = df_all_data.attrs['cvp_options']

    return render_template('index.html',
        revenue_years=revenue_years,
        market_size_years=market_size_years,
        svp_options=svp_options,
        cvp_options=cvp_options,
        regions=regions,
        default_display_start=default_display_start,
        default_display_end=default_display_end,
        default_cagr_start=default_cagr_start,
        default_cagr_end=default_cagr_end,
        default_pred_start=default_pred_start,
        warnings=warnings,
    )


@app.route('/results_data', methods=['POST'])
def results_data():
    cagr_start_year    = int(request.form.get('cagr_start_year'))
    cagr_end_year      = int(request.form.get('cagr_end_year'))
    display_start_year = int(request.form.get('display_start_year'))
    display_end_year   = int(request.form.get('display_end_year'))
    pred_start_year    = int(request.form.get('pred_start_year'))
    selected_svps      = request.form.getlist('svp')
    selected_cvps      = request.form.getlist('cvp')
    region             = request.form.get('region')
    use_custom_cagr    = request.form.get('use_custom_cagr', 'no')
    subseg_filter      = request.form.get('subseg_filter', 'all')
    aggregation_level  = request.form.get('aggregation_level', 'offer')

    selected_columns_json = request.form.get('selected_columns', '[]')
    try:
        selected_columns = json.loads(selected_columns_json)
    except Exception:
        selected_columns = None

    result = resultat_data(
        df_all_data=df_all_data,
        display_start_year=display_start_year,
        display_end_year=display_end_year,
        cagr_start_year=cagr_start_year,
        cagr_end_year=cagr_end_year,
        pred_start_year=pred_start_year,
        selected_svps=selected_svps,
        region=region,
        use_custom_cagr=use_custom_cagr,
        selected_columns=selected_columns,
        subseg_filter=subseg_filter,
        aggregation_level=aggregation_level,
        selected_cvps=selected_cvps,
    )
    return jsonify(result)


@app.route('/download_excel', methods=['POST'])
def download_excel():
    try:
        data        = request.json
        table_data  = data.get('data', [])
        columns     = data.get('columns', [])
        disp_start  = data.get('display_start_year', '')
        disp_end    = data.get('display_end_year', '')

        if not table_data or not columns:
            return jsonify({'error': 'No data to export'}), 400

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Analysis {disp_start}-{disp_end}"

        hdr_fill  = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        pred_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        pred_bl_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
        real_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        tot_fill  = PatternFill(start_color="E7E7FF", end_color="E7E7FF", fill_type="solid")
        tot_font  = Font(bold=True, color="667EEA", size=10)
        border    = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'))

        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for ri, row_data in enumerate(table_data, 2):
            is_total = row_data.get('Offer') == 'TOTAL'
            for ci, col in enumerate(columns, 1):
                value = row_data.get(col, '')
                cell  = ws.cell(row=ri, column=ci, value=value)
                cell.border = border
                if is_total:
                    cell.fill = tot_fill; cell.font = tot_font
                elif col.startswith('Predicted_BL_'):
                    cell.fill = pred_bl_fill
                elif col.startswith('Predicted_'):
                    cell.fill = pred_fill
                elif col.startswith('Real_'):
                    cell.fill = real_fill
                if col.startswith(('Revenue_','Real_','Predicted_')):
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif 'CAGR' in col:
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')

        for ci, col in enumerate(columns, 1):
            mx = max((len(str(ws.cell(row=r, column=ci).value or ''))
                      for r in range(1, len(table_data)+2)), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 2, 50)

        ws.freeze_panes = 'A2'
        out = BytesIO(); wb.save(out); out.seek(0)
        return send_file(out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'SVP_Analysis_{disp_start}_{disp_end}.xlsx')

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/files')
def files():
    return render_template('files.html')


# Mirrors the `requiredColumns` in FILE_TYPES (files.html) — the "important"
# columns kept when exporting a tracked source file.
REQUIRED_COLUMNS = {
    'cagr_svp': ['catégorie', 'pr_offer', 'pr_offer_line', 'pr_sub_domain',
                 'pr_domain', 'pr_sub_business_line', 'pr_business_line'],
    'market': ['Segment', 'Sub-segment 1', 'Sub-segment 2', 'Sub-segment 3',
               'Million EUR', 'Year'],
    'all_products': ['rep_code', 'Business_Line', 'Sub_Business_Line', 'Domain',
                      'Sub_Domain', 'offer_line', 'offer', 'Product',
                      'Strategic (for SVPs)', 'CVP', 'Delivery_Zone (Region)'],
    'data_fr': ['Rep_Code', 'Offer', 'Period', '_S_Revenue_actual'],
    'data_intl': ['Rep_Code', 'Offer', 'Period', '_S_Revenue_actual'],
}


@app.route('/download_file/<file_type>')
def download_file(file_type):
    try:
        if file_type not in REQUIRED_COLUMNS:
            return jsonify({'error': 'Invalid file type'}), 400

        path = dict(TRACKED_FILES).get(file_type)
        if not path or not os.path.exists(path):
            return jsonify({'error': 'File not found'}), 404

        # Read the file the same way it's read/validated on upload.
        if file_type == 'cagr_svp':
            xls = pd.ExcelFile(path)
            relevant_sheets = [s for s in xls.sheet_names
                                if not s.startswith('MIF') and 'GLOBAL' not in s]
            df = None
            for sheet in relevant_sheets:
                df_sheet = pd.read_excel(path, sheet_name=sheet, skiprows=2)
                if 'catégorie' in df_sheet.columns:
                    df = df_sheet
                    break
            if df is None:
                return jsonify({'error': "Missing 'catégorie' column in CAGR SVP sheets"}), 400

        elif file_type == 'market':
            sheet_names = pd.ExcelFile(path).sheet_names
            if 'DATA BASE MARKET FORECAST' in sheet_names:
                df = pd.read_excel(path, sheet_name='DATA BASE MARKET FORECAST', skiprows=5)
            else:
                df = pd.read_excel(path)

        elif file_type == 'all_products':
            df = pd.read_excel(path)
            df.columns = df.columns.str.strip()

        else:  # data_fr, data_intl
            df = pd.read_excel(path)

        # Keep only the "important" columns, matched case-insensitively,
        # in the order defined by REQUIRED_COLUMNS.
        col_lookup = {str(c).strip().lower(): c for c in df.columns}
        wanted = REQUIRED_COLUMNS[file_type]
        selected_cols = [col_lookup[w.lower()] for w in wanted if w.lower() in col_lookup]
        missing_cols  = [w for w in wanted if w.lower() not in col_lookup]

        if not selected_cols:
            return jsonify({'error': 'None of the required columns were found in this file'}), 400

        df_filtered = df[selected_cols]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = file_type[:31]

        hdr_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        border   = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'))

        for ci, col in enumerate(selected_cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for ri, (_, row) in enumerate(df_filtered.iterrows(), 2):
            for ci, col in enumerate(selected_cols, 1):
                value = row[col]
                if pd.isna(value):
                    value = ''
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.border = border

        for ci, col in enumerate(selected_cols, 1):
            mx = max((len(str(ws.cell(row=r, column=ci).value or ''))
                      for r in range(1, len(df_filtered) + 2)), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 2, 50)

        ws.freeze_panes = 'A2'

        if missing_cols:
            note_ws = wb.create_sheet('Notes')
            note_ws.cell(row=1, column=1,
                         value=f"Columns not found in source file: {', '.join(missing_cols)}")

        out = BytesIO(); wb.save(out); out.seek(0)
        return send_file(out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{file_type}_important_columns.xlsx')

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/settings')
def settings():
    return render_template('settings.html')


TRACKED_FILES = [
    ('cagr_svp',    DATA_CAGR_SVP),
    ('market',      DATA_MARKET_SIZE),
    ('all_products', DATA_ALL_PRODUCTS),
    ('market_hierarchies', DATA_MARKET_HIERARCHIES),
    ('data_fr',     DATA_REVENUE_FR),
    ('data_intl',   DATA_REVENUE_INTL),
]


@app.route('/file_status')
def file_status():
    try:
        files_info = {}
        warnings   = []
        for key, path in TRACKED_FILES:
            if os.path.exists(path):
                stat     = os.stat(path)
                mod_time = datetime.fromtimestamp(stat.st_mtime)
                days_old = (datetime.now() - mod_time).days
                try:
                    if key == 'cagr_svp':
                        # Multi-sheet workbook: count rows across every
                        # relevant (non MIF*/GLOBAL*) sheet.
                        xls  = pd.ExcelFile(path)
                        rows = sum(
                            len(pd.read_excel(path, sheet_name=s, skiprows=2))
                            for s in xls.sheet_names
                            if not s.startswith('MIF') and 'GLOBAL' not in s
                        )
                    else:
                        rows = len(pd.read_excel(path))
                    files_info[key] = {
                        'modified': mod_time.strftime('%Y-%m-%d %H:%M'),
                        'rows': rows,
                        'size': f"{stat.st_size / 1024:.1f} KB",
                        'days_old': days_old
                    }
                    if days_old > 30:
                        warnings.append(f"⚠️ {key} file is {days_old} days old.")
                except Exception:
                    pass
        return jsonify({'files': files_info, 'warnings': warnings})
    except Exception as e:
        return jsonify({'files': {}, 'warnings': [str(e)]}), 500


@app.route('/file_history')
def file_history():
    try:
        history = []
        backup_dirs = {
            file_type: os.path.join(DATA_DIR, 'backups', file_type)
            for file_type, _ in TRACKED_FILES
        }
        for file_type, backup_dir in backup_dirs.items():
            if os.path.exists(backup_dir):
                for filename in sorted(os.listdir(backup_dir), reverse=True)[:5]:
                    filepath = os.path.join(backup_dir, filename)
                    if os.path.isfile(filepath):
                        try:
                            df       = pd.read_excel(filepath, sheet_name=0)
                            stat     = os.stat(filepath)
                            mod_time = datetime.fromtimestamp(stat.st_mtime)
                            history.append({
                                'file_type': file_type,
                                'timestamp': mod_time.strftime('%Y-%m-%d %H:%M:%S'),
                                'rows': len(df),
                                'filename': filename
                            })
                        except Exception:
                            pass
        history = sorted(history, key=lambda x: x['timestamp'], reverse=True)
        return jsonify({'history': history})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/upload_custom_cagr', methods=['POST'])
def upload_custom_cagr():
    try:
        uploaded_file = request.files.get('file')
        if not uploaded_file:
            return jsonify({'error': 'No file provided'}), 400
        file_bytes = uploaded_file.read()
        uploaded_file.seek(0)
        try:
            df = pd.read_excel(uploaded_file)
        except Exception:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file)
        df.columns = df.columns.str.strip()
        offer_col = next((c for c in df.columns if c.lower() == 'offer'), None)
        cagr_col  = next((c for c in df.columns if c.lower() == 'cagr'),  None)
        if not cagr_col:
            # Also accept a 'CAGR 20XX/20YY' style header (the BL CAGR's
            # own period); pipeline.load_custom_cagr() extracts the years
            # from this and renames it back to 'CAGR' when reading it.
            cagr_col = next(
                (c for c in df.columns
                 if re.search(r'cagr.*?\d{4}\s*/\s*\d{4}', c, re.IGNORECASE)),
                None
            )
        if not offer_col or not cagr_col:
            return jsonify({'error': f'Missing columns. Found: {", ".join(df.columns)}'}), 400
        valid_count = sum(
            1 for _, r in df.iterrows()
            if not pd.isna(r[offer_col]) and not pd.isna(r[cagr_col])
            and r[offer_col] != '' and r[cagr_col] != ''
            and _is_numeric(r[cagr_col])
        )
        if valid_count == 0:
            return jsonify({'error': 'No valid entries found.'}), 400
        custom_cagr_path = os.path.join(DATA_DIR, 'custom_cagr.xlsx')
        if os.path.exists(custom_cagr_path):
            backup_old_file('custom_cagr', custom_cagr_path)
        with open(custom_cagr_path, 'wb') as f:
            f.write(file_bytes)
        return jsonify({'success': True, 'message': f'Uploaded ({valid_count} valid entries)',
                        'rows': valid_count, 'timestamp': datetime.now().isoformat()})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _is_numeric(v):
    try:
        float(v); return True
    except (ValueError, TypeError):
        return False


@app.route('/custom_cagr_status')
def custom_cagr_status():
    try:
        filepath = os.path.join(DATA_DIR, 'custom_cagr.xlsx')
        if os.path.exists(filepath):
            df       = pd.read_excel(filepath)
            stat     = os.stat(filepath)
            mod_time = datetime.fromtimestamp(stat.st_mtime)
            return jsonify({'exists': True,
                            'modified': mod_time.strftime('%Y-%m-%d %H:%M'),
                            'rows': len(df),
                            'size': f"{stat.st_size/1024:.1f} KB"})
        return jsonify({'exists': False})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/upload_file', methods=['POST'])
def upload_file():
    try:
        file_type     = request.form.get('file_type')
        uploaded_file = request.files.get('file')
        if not uploaded_file: return jsonify({'error': 'No file provided'}), 400
        if not file_type:     return jsonify({'error': 'File type not specified'}), 400
        file_bytes = uploaded_file.read()
        uploaded_file.seek(0)

        if file_type == 'cagr_svp':
            # Multi-sheet workbook: category hierarchy source for the new
            # Rep_Code → catégorie → ID mapper. Validate that at least one
            # relevant (non MIF*/GLOBAL*) sheet has a 'catégorie' column
            # once the 2 header rows are skipped.
            try:
                xls = pd.ExcelFile(uploaded_file)
            except Exception:
                return jsonify({'error': 'Could not read workbook'}), 400
            relevant_sheets = [s for s in xls.sheet_names
                              if not s.startswith("MIF") and "GLOBAL" not in s]
            if not relevant_sheets:
                return jsonify({'error': 'No relevant sheets found (all sheets are MIF/GLOBAL)'}), 400
            df = None
            for sheet in relevant_sheets:
                uploaded_file.seek(0)
                df_sheet = pd.read_excel(uploaded_file, sheet_name=sheet, skiprows=2)
                if 'catégorie' in df_sheet.columns:
                    df = df_sheet
                    break
            if df is None:
                return jsonify({'error': "Missing 'catégorie' column in CAGR SVP sheets"}), 400
            target_path = DATA_CAGR_SVP

        elif file_type == 'market':
            uploaded_file.seek(0)
            sheet_names = pd.ExcelFile(uploaded_file).sheet_names
            uploaded_file.seek(0)
            if 'DATA BASE MARKET FORECAST' in sheet_names:
                df = pd.read_excel(uploaded_file, sheet_name='DATA BASE MARKET FORECAST', skiprows=5)
            else:
                df = pd.read_excel(uploaded_file)
            # 'ID' is no longer required — Million EUR is now joined on the
            # (Segment, Sub-segment 1-3, Region) hierarchy directly instead
            # of a Market Sizing 'ID' column (see mapping.py).
            required = ['Year','Million EUR','Region','Segment']
            missing  = [c for c in required if c not in df.columns]
            if missing: return jsonify({'error': f'Missing: {", ".join(missing)}'}), 400
            target_path = DATA_MARKET_SIZE

        elif file_type == 'all_products':
            df = pd.read_excel(uploaded_file)
            df.columns = df.columns.str.strip()
            cols_lower = [c.lower() for c in df.columns]
            required   = ['rep_code', 'offer']
            missing    = [c for c in required if c not in cols_lower]
            if missing: return jsonify({'error': f'Missing: {", ".join(missing)}'}), 400
            target_path = DATA_ALL_PRODUCTS

        elif file_type == 'market_hierarchies':
            # Pre-identified catégorie -> Segment/Sub-segment lookup table,
            # used directly (exact-match join, no semantic/fuzzy matching)
            # to fill in each Rep_Code's Segment/Sub-segment hierarchy.
            try:
                df = pd.read_excel(uploaded_file, sheet_name='Sheet1')
            except Exception:
                return jsonify({'error': "Could not read 'Sheet1' from workbook"}), 400
            required = ['catégorie (ID)', 'Segment_FR', 'Sub-segment 1_FR',
                        'Sub-segment 2_FR', 'Sub-segment 3_FR']
            missing  = [c for c in required if c not in df.columns]
            if missing: return jsonify({'error': f'Missing: {", ".join(missing)}'}), 400
            target_path = DATA_MARKET_HIERARCHIES

        elif file_type in ('data_fr', 'data_intl'):
            df = pd.read_excel(uploaded_file)
            required = ['Offer','Period','_S_Revenue_actual']
            missing  = [c for c in required if c not in df.columns]
            if missing: return jsonify({'error': f'Missing: {", ".join(missing)}'}), 400
            target_path = DATA_REVENUE_FR if file_type == 'data_fr' else DATA_REVENUE_INTL

        else:
            return jsonify({'error': 'Invalid file type'}), 400

        if os.path.exists(target_path):
            backup_old_file(file_type, target_path)
        with open(target_path, 'wb') as f:
            f.write(file_bytes)

        # Regenerate the combined data.xlsx whenever a regional revenue
        # file changes.
        if file_type in ('data_fr', 'data_intl'):
            combine_revenue_files()

        # Rebuild the Rep_Code/catégorie/Segment mapper whenever one of
        # its source files changes.
        if file_type in ('cagr_svp', 'all_products', 'market_hierarchies'):
            build_category_mapper(force_rebuild=True)

        global revenue_years, market_size_years, df_all_data
        revenue_years     = load_revenue_data()
        market_size_years = load_market_size_years()
        df_all_data       = load_all_data()
        return jsonify({'success': True, 'message': f'{file_type} uploaded',
                        'rows': len(df), 'timestamp': datetime.now().isoformat()})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def backup_old_file(file_type, source_path):
    try:
        backup_dir = os.path.join(DATA_DIR, 'backups', file_type)
        os.makedirs(backup_dir, exist_ok=True)
        timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(backup_dir, f'backup_{timestamp}.xlsx')
        if os.path.exists(source_path):
            shutil.copy2(source_path, backup_path)
    except Exception as e:
        print(f"⚠ Backup failed: {e}")


@app.route('/back')
def back():
    return redirect('/')


if __name__ == '__main__':
    app.run(debug=True)